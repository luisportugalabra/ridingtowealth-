#!/usr/bin/env python3
"""
Reads portfolio.xlsx and updates portfolio-positions.json with latest returns.
Run locally: python3 scripts/update-from-excel.py
Then: git add portfolio-positions.json && git commit -m "Update returns" && git push
"""
import json, os, sys

POSSIBLE_PATHS = [
    os.path.expanduser("~/Desktop/claude/portfolio/portfolio.xlsx"),
    os.path.expanduser("~/Desktop/stocks/portfolio.xlsx"),
]

xlsx_path = None
for p in POSSIBLE_PATHS:
    if os.path.exists(p):
        xlsx_path = p
        break

if not xlsx_path:
    print("ERROR: portfolio.xlsx not found")
    sys.exit(1)

print(f"Reading: {xlsx_path}")

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb['DASHBOARD']

# Row 17: 2024, Row 18: 2025, Row 19: 2026 YTD (excl. cash)
# Columns: A=Year, B=GERAL EUR, C=GERAL USD, D=VC2 EUR, E=VC2 USD, F=VC2T EUR, G=VC2T USD, H=S&P 500
def pct(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return round(v * 100, 1) if abs(v) < 2 else round(v, 1)
    return None

r17 = [ws.cell(row=17, column=c).value for c in range(1, 9)]  # 2024
r18 = [ws.cell(row=18, column=c).value for c in range(1, 9)]  # 2025
r19 = [ws.cell(row=19, column=c).value for c in range(1, 9)]  # 2026 YTD

print(f"\n2024:     GERAL USD={pct(r17[2])}%  VC2 USD={pct(r17[4])}%  VC2T USD={pct(r17[6])}%  S&P={pct(r17[7])}%")
print(f"2025:     GERAL USD={pct(r18[2])}%  VC2 USD={pct(r18[4])}%  VC2T USD={pct(r18[6])}%  S&P={pct(r18[7])}%")
print(f"2026 YTD: GERAL USD={pct(r19[2])}%  VC2 USD={pct(r19[4])}%  VC2T USD={pct(r19[6])}%  S&P={pct(r19[7])}%")

# Load positions JSON
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
pos_path = os.path.join(ROOT_DIR, "portfolio-positions.json")

with open(pos_path) as f:
    positions = json.load(f)

# Update returns (USD)
positions["annual_returns"] = {
    "discretionary": {},
    "vc2": {},
    "vc2trend": {}
}

# 2024
if pct(r17[2]): positions["annual_returns"]["discretionary"]["2024"] = pct(r17[2])
if pct(r17[4]): positions["annual_returns"]["vc2"]["2024"] = pct(r17[4])
if pct(r17[6]): positions["annual_returns"]["vc2trend"]["2024"] = pct(r17[6])

# 2025
if pct(r18[2]): positions["annual_returns"]["discretionary"]["2025"] = pct(r18[2])
if pct(r18[4]): positions["annual_returns"]["vc2"]["2025"] = pct(r18[4])
if pct(r18[6]): positions["annual_returns"]["vc2trend"]["2025"] = pct(r18[6])

# 2026 YTD
if pct(r19[2]): positions["annual_returns"]["discretionary"]["2026_ytd"] = pct(r19[2])
if pct(r19[4]): positions["annual_returns"]["vc2"]["2026_ytd"] = pct(r19[4])
if pct(r19[6]): positions["annual_returns"]["vc2trend"]["2026_ytd"] = pct(r19[6])

# S&P 500
positions["sp500_returns"] = {}
if pct(r17[7]): positions["sp500_returns"]["2024"] = pct(r17[7])
if pct(r18[7]): positions["sp500_returns"]["2025"] = pct(r18[7])
if pct(r19[7]): positions["sp500_returns"]["2026_ytd"] = pct(r19[7])

# Save
with open(pos_path, "w") as f:
    json.dump(positions, f, indent=2)

print(f"\nSaved to {pos_path}")
print(f"\nValues for site (USD):")
for strat, rets in positions["annual_returns"].items():
    print(f"  {strat}: {rets}")
print(f"  S&P 500: {positions['sp500_returns']}")
