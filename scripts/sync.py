#!/usr/bin/env python3
"""
ONE COMMAND TO SYNC EVERYTHING.

Usage: python3 ~/bivarcapital/scripts/sync.py

Reads portfolio.xlsx → saves positions + returns + sync baseline → fetches prices → pushes to GitHub.

The GitHub Action (daily) then uses the sync baseline to calculate live YTD:
  YTD_live = (1 + sync_ytd) × (value_today / sync_value) - 1
"""
import json, os, sys, subprocess
from datetime import datetime, timezone

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
os.chdir(ROOT_DIR)

# Find Excel
EXCEL_PATHS = [
    os.path.expanduser("~/Desktop/claude/portfolio/portfolio.xlsx"),
    os.path.expanduser("~/Desktop/stocks/portfolio.xlsx"),
]
xlsx_path = next((p for p in EXCEL_PATHS if os.path.exists(p)), None)
if not xlsx_path:
    print("ERROR: portfolio.xlsx not found")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl

try:
    import yfinance as yf
except ImportError:
    os.system("pip install yfinance -q")
    import yfinance as yf

print(f"📂 Reading: {xlsx_path}")

# ============ 1. READ POSITIONS ============
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb['POSITIONS']

STRATEGY_MAP = {"GERAL": "discretionary", "VC2": "vc2", "VC2 TREND": "vc2trend"}
positions = {"discretionary": [], "vc2": [], "vc2trend": []}

for row in ws.iter_rows(min_row=13, max_row=200, values_only=True):
    ticker, broker, strategy, qty, avg_cost = row[0], row[1], row[2], row[3], row[4]
    ccy = row[10] if len(row) > 10 else "USD"
    if not ticker or not strategy or not qty:
        continue
    if str(ticker).startswith(('TOTAL', 'IBKR', 'TastyTrade', 'GRAND', '—')):
        continue
    strat_key = STRATEGY_MAP.get(str(strategy).strip())
    if not strat_key:
        continue
    t = str(ticker).strip()
    if t == "BRK.B": t = "BRK-B"
    positions[strat_key].append({"ticker": t, "qty": float(qty), "avg_cost": float(avg_cost), "ccy": str(ccy).strip()})

total_pos = sum(len(v) for v in positions.values())
print(f"📊 Positions: {total_pos}")

# ============ 2. READ RETURNS FROM EXCEL ============
ws_dash = wb['DASHBOARD']
def pct(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return round(v * 100, 1) if abs(v) < 2 else round(v, 1)
    return None

r18 = [ws_dash.cell(row=18, column=c).value for c in range(1, 9)]
r19 = [ws_dash.cell(row=19, column=c).value for c in range(1, 9)]
# Rows 22-23: Discretionary "incl. cash" (USD) — the correct values
r22 = [ws_dash.cell(row=22, column=c).value for c in range(1, 9)]
r23 = [ws_dash.cell(row=23, column=c).value for c in range(1, 9)]

returns = {
    "discretionary": {"2025": pct(r22[2]), "2026_ytd": pct(r23[2])},
    "vc2": {"2025": pct(r18[4]), "2026_ytd": pct(r19[4])},
    "vc2trend": {"2025": pct(r18[6]), "2026_ytd": pct(r19[6])},
}
sp500 = {"2025": pct(r18[7]), "2026_ytd": pct(r19[7])}

# Clean None values
for strat in returns.values():
    for k in list(strat.keys()):
        if strat[k] is None: del strat[k]

print(f"📈 YTD from Excel: Disc={returns['discretionary'].get('2026_ytd','?')}%, VC2={returns['vc2'].get('2026_ytd','?')}%, VC2T={returns['vc2trend'].get('2026_ytd','?')}%")

# ============ 3. FETCH LIVE PRICES ============
YF_MAP = {
    "BRK-B": "BRK-B", "2330": "2330.TW", "WISE": "WISE.L",
    "CLNX": "CLNX.MC", "PRX": "PRX.AS", "EXO": "EXO.MI",
}

all_tickers = set()
for strat in positions.values():
    for p in strat:
        all_tickers.add(YF_MAP.get(p["ticker"], p["ticker"]))

print(f"🔄 Fetching {len(all_tickers)} prices...")
data = yf.download(list(all_tickers), period="1d", progress=False)
prices = {}
for t in all_tickers:
    try:
        close = data["Close"][t].dropna()
        if len(close) > 0: prices[t] = float(close.iloc[-1])
    except: pass

fx_data = yf.download(["EURUSD=X", "GBPUSD=X"], period="1d", progress=False)
eur_usd = float(fx_data["Close"]["EURUSD=X"].dropna().iloc[-1]) if "EURUSD=X" in fx_data["Close"] else 1.13
gbp_usd = float(fx_data["Close"]["GBPUSD=X"].dropna().iloc[-1]) if "GBPUSD=X" in fx_data["Close"] else 1.27
twd_usd = 1.0 / 32.0

print(f"💰 Got {len(prices)} prices")

# ============ 4. BUILD PORTFOLIO ============
def price_usd(ticker, ccy):
    yf_t = YF_MAP.get(ticker, ticker)
    p = prices.get(yf_t)
    if p is None: return None
    if ccy == "EUR": return p * eur_usd
    if ccy == "GBP": return (p / 100) * gbp_usd
    if ccy == "TWD": return p * twd_usd
    return p

now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

output = {
    "updated": now,
    "strategies": {},
    "returns": returns,
    "sp500_returns": sp500,
    "sync": {}
}

for strat_name, strat_pos in positions.items():
    sd = {"positions": [], "total_value": 0}
    for p in strat_pos:
        pu = price_usd(p["ticker"], p["ccy"])
        if pu is None:
            pu = p["avg_cost"]
            print(f"  ⚠️  No price for {p['ticker']}")
        mv = pu * p["qty"]
        sd["positions"].append({"ticker": p["ticker"], "qty": p["qty"], "price": round(pu, 2), "mkt_value": round(mv, 2), "weight": 0})
        sd["total_value"] += mv

    for pos in sd["positions"]:
        pos["weight"] = round((pos["mkt_value"] / sd["total_value"]) * 100, 1) if sd["total_value"] > 0 else 0
    sd["positions"].sort(key=lambda x: x["weight"], reverse=True)
    sd["total_value"] = round(sd["total_value"], 2)
    sd["n_positions"] = len(sd["positions"])
    output["strategies"][strat_name] = sd

    # Save sync baseline for this strategy
    # YTD_live = (1 + sync_ytd) × (value_today / sync_value) - 1
    sync_ytd = returns.get(strat_name, {}).get("2026_ytd", 0)
    output["sync"][strat_name] = {
        "date": now,
        "value": round(sd["total_value"], 2),
        "ytd_at_sync": sync_ytd
    }

# S&P 500 sync
output["sync"]["sp500"] = {
    "date": now,
    "ytd_at_sync": sp500.get("2026_ytd", 0)
}

# ============ 5. GENERATE ACTIVITY LOG ============
STRAT_NAMES = {"discretionary": "Discretionary", "vc2": "Value Strategy", "vc2trend": "Momentum + Value"}
activity = []
date_short = datetime.now(timezone.utc).strftime("%b %-d")

# Load old data to compare
try:
    with open("portfolio-data.json") as f:
        old_data = json.load(f)
    old_activity = old_data.get("activity", [])
except:
    old_data = {}
    old_activity = []

for strat_name in ["discretionary", "vc2", "vc2trend"]:
    old_tickers = set()
    if "strategies" in old_data and strat_name in old_data["strategies"]:
        old_tickers = {p["ticker"] for p in old_data["strategies"][strat_name]["positions"]}
    new_tickers = {p["ticker"] for p in output["strategies"][strat_name]["positions"]}

    added = new_tickers - old_tickers
    removed = old_tickers - new_tickers
    sname = STRAT_NAMES[strat_name]

    if added and removed:
        activity.append({"date": date_short, "text": f"{sname} rebalanced \u2014 {len(added)} added, {len(removed)} removed"})
    elif added:
        activity.append({"date": date_short, "text": f"{sname} \u2014 added {', '.join(sorted(added))}"})
    elif removed:
        activity.append({"date": date_short, "text": f"{sname} \u2014 removed {', '.join(sorted(removed))}"})

if not activity:
    activity.append({"date": date_short, "text": "Portfolio synced \u2014 positions updated"})

# Keep last 5 entries (new + old)
output["activity"] = (activity + old_activity)[:5]

# ============ 6. SAVE ============
with open("portfolio-data.json", "w") as f:
    json.dump(output, f, indent=2)

print(f"\n✅ portfolio-data.json saved")
print(f"   Sync baselines:")
for s, sv in output["sync"].items():
    val = sv.get('value', 0)
    print(f"   {s}: value=${val:,.0f} ytd={sv['ytd_at_sync']}%") if isinstance(val, (int, float)) else print(f"   {s}: ytd={sv['ytd_at_sync']}%")

# ============ 6. GIT PUSH ============
print(f"\n🚀 Pushing...")
subprocess.run(["git", "add", "portfolio-data.json"], check=True)
result = subprocess.run(["git", "diff", "--staged", "--quiet"])
if result.returncode != 0:
    subprocess.run(["git", "commit", "-m", f"Sync portfolio {now[:10]}"], check=True)
    subprocess.run(["git", "push"], check=True)
    print("✅ Site updated!")
else:
    print("ℹ️  No changes")
