#!/usr/bin/env python3
"""
VC2, Trending Value & Momentum Portfolio Generator
Based on James O'Shaughnessy's "What Works on Wall Street"
and Novy-Marx (2012) Intermediate Momentum

Auto-detects the most recent "vc2 world_*.csv" file in this folder.
Generates one Excel file with:
  - VC2 Full Ranking (all stocks, including financials)
  - TV50 for each filter (no financials, positive P/E only)
  - MOM50 for each filter (intermediate momentum, all sectors)
And wizard-data.json for the website.
"""

import pandas as pd
import numpy as np
import glob
import os
import sys
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIG
# ============================================================================
FILTERS = {
    'Global': None,
    'USA': ['United States'],
    'Europe': [
        'United Kingdom', 'France', 'Germany', 'Switzerland', 'Netherlands',
        'Sweden', 'Norway', 'Denmark', 'Finland', 'Spain', 'Italy', 'Belgium',
        'Austria', 'Ireland', 'Portugal', 'Greece', 'Luxembourg', 'Poland',
        'Czech Republic', 'Hungary', 'Romania', 'Iceland', 'Estonia',
        'Lithuania', 'Slovenia', 'Cyprus', 'Monaco',
    ],
    'Asia': [
        'Japan', 'China', 'South Korea', 'Taiwan', 'Hong Kong', 'Singapore',
        'India', 'Thailand', 'Malaysia', 'Indonesia', 'Philippines', 'Vietnam',
        'Bangladesh', 'Sri Lanka', 'Pakistan',
    ],
    'Emerging': [
        'Brazil', 'Mexico', 'Chile', 'Colombia', 'Peru', 'Argentina',
        'China', 'India', 'Indonesia', 'Thailand', 'Malaysia', 'Philippines',
        'Vietnam', 'Bangladesh', 'Sri Lanka', 'Pakistan',
        'South Africa', 'Nigeria', 'Kenya', 'Egypt',
        'Turkey', 'Saudi Arabia', 'United Arab Emirates', 'Qatar', 'Kuwait',
        'Bahrain', 'Russian Federation',
        'Poland', 'Czech Republic', 'Hungary', 'Romania',
    ],
}

TV_SIZE = 50
MOM_SIZE = 50
JSON_TOP_N = 200      # For value/trend (filtered to decile 1, good buffer for mkt cap filtering)
MOM_JSON_TOP_N = 500  # For momentum (screen entire market, need large pool for mkt cap filtering)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
COUNTRY_COL = 'Country or region of registration'

# ============================================================================
# 1. AUTO-DETECT CSV
# ============================================================================
csv_files = glob.glob(os.path.join(SCRIPT_DIR, "vc2 world_*.csv"))
if not csv_files:
    csv_files = glob.glob(os.path.join(SCRIPT_DIR, "vc2*.csv"))
if not csv_files:
    print("ERRO: Nenhum ficheiro CSV encontrado.")
    sys.exit(1)

csv_path = max(csv_files, key=os.path.getmtime)
csv_name = os.path.basename(csv_path)
print(f"CSV: {csv_name}")

file_date = ""
try:
    date_part = csv_name.split("_")[1].replace(".csv", "").split("_")[0]
    datetime.strptime(date_part, "%Y-%m-%d")
    file_date = date_part
except:
    file_date = datetime.now().strftime("%Y-%m-%d")

out_path = os.path.join(SCRIPT_DIR, f"VC2_Trending_Value_{file_date}.xlsx")

# ============================================================================
# 2. LOAD DATA
# ============================================================================
print("A carregar dados...")
df = pd.read_csv(csv_path)
total_raw = len(df)
print(f"  {total_raw} ações")

# ============================================================================
# 3. PERCENTILE FUNCTIONS
# ============================================================================
def percentile_inc(series, x):
    if pd.isna(x):
        return 0.5
    arr = series.dropna()
    n = len(arr)
    if n < 2:
        return 0.5
    return (arr < x).sum() / (n - 1)

def percentile_shareholder(series, x):
    if pd.isna(x):
        x = 0
    arr = series.fillna(0)
    n = len(arr)
    if n < 2:
        return 0.5
    return 1 - ((arr < x).sum() / (n - 1))

# ============================================================================
# 4. CALCULATE VC2 (on ALL stocks for Full Ranking)
# ============================================================================
print("A calcular VC2...")
df['Shareholder Yield'] = df['Buyback yield %'].fillna(0) + df['Dividend yield %, Trailing 12 months'].fillna(0)

valuation_cols = [
    'Price to earnings ratio',
    'Price to sales ratio',
    'Price to book ratio',
    'Price to cash flow ratio',
    'Enterprise value to EBITDA ratio, Trailing 12 months'
]

for col in valuation_cols:
    df[f'{col} pct'] = df[col].apply(lambda x: percentile_inc(df[col], x))

df['Shareholder Yield pct'] = df['Shareholder Yield'].apply(
    lambda x: percentile_shareholder(df['Shareholder Yield'], x)
)

pct_cols = [f'{c} pct' for c in valuation_cols] + ['Shareholder Yield pct']
df['VC2'] = df[pct_cols].sum(axis=1)
df['VC2 Decile'] = pd.qcut(df['VC2'], 10, labels=False, duplicates='drop') + 1

# ============================================================================
# 4b. CALCULATE INTERMEDIATE MOMENTUM (12M - 6M, Novy-Marx 2012)
# ============================================================================
mom6_col = 'Performance % 6 months'
mom12_col = 'Performance % 1 year'

has_momentum = mom12_col in df.columns
if has_momentum:
    df['Intermediate Momentum'] = df[mom12_col] - df[mom6_col]
    print(f"  Intermediate Momentum calculado (12M - 6M)")
else:
    print(f"  AVISO: Coluna '{mom12_col}' não encontrada — momentum não será gerado")

# ============================================================================
# 4c. FILTERED DF for Value/Trend (no financials, positive P/E)
# ============================================================================
df_filtered = df[df['Sector'] != 'Finance'].copy()
df_filtered = df_filtered[df_filtered['Price to earnings ratio'].notna() & (df_filtered['Price to earnings ratio'] > 0)].copy()
print(f"  Filtro Value/Trend (sem Finance, P/E positivo): {len(df_filtered)} ações (removidas {total_raw - len(df_filtered)})")

# Recalculate VC2 on filtered data for proper deciles
for col in valuation_cols:
    df_filtered[f'{col} pct'] = df_filtered[col].apply(lambda x: percentile_inc(df_filtered[col], x))
df_filtered['Shareholder Yield pct'] = df_filtered['Shareholder Yield'].apply(
    lambda x: percentile_shareholder(df_filtered['Shareholder Yield'], x)
)
df_filtered['VC2'] = df_filtered[pct_cols].sum(axis=1)
df_filtered['VC2 Decile'] = pd.qcut(df_filtered['VC2'], 10, labels=False, duplicates='drop') + 1

cheapest_filtered = df_filtered[df_filtered['VC2 Decile'] == 1].copy()
print(f"  VC2 Decil 1 (filtrado): {len(cheapest_filtered)} ações")

# ============================================================================
# 5. BUILD EXCEL SHEETS
# ============================================================================
output_cols = [
    'Symbol', 'Description', COUNTRY_COL, 'Sector',
    'Market capitalization',
    'Price to earnings ratio', 'Price to sales ratio', 'Price to book ratio',
    'Price to cash flow ratio', 'Enterprise value to EBITDA ratio, Trailing 12 months',
    'Buyback yield %', 'Dividend yield %, Trailing 12 months', 'Shareholder Yield',
    'VC2', 'VC2 Decile',
    'Performance % 6 months',
    'Return on invested capital %, Trailing 12 months'
]

mom_output_cols = [
    'Symbol', 'Description', COUNTRY_COL, 'Sector',
    'Market capitalization',
    'Performance % 1 year', 'Performance % 6 months', 'Intermediate Momentum',
    'Return on invested capital %, Trailing 12 months'
]

header_map = {
    'Symbol': 'Ticker', 'Description': 'Company',
    COUNTRY_COL: 'Country', 'Sector': 'Sector',
    'Market capitalization': 'Mkt Cap ($M)',
    'Price to earnings ratio': 'P/E', 'Price to sales ratio': 'P/S',
    'Price to book ratio': 'P/B', 'Price to cash flow ratio': 'P/CF',
    'Enterprise value to EBITDA ratio, Trailing 12 months': 'EV/EBITDA',
    'Buyback yield %': 'Buyback Yield %',
    'Dividend yield %, Trailing 12 months': 'Div Yield %',
    'Shareholder Yield': 'Shareholder Yield %',
    'VC2': 'VC2 Score', 'VC2 Decile': 'VC2 Decile',
    'Performance % 6 months': 'Momentum 6M %',
    'Performance % 1 year': 'Momentum 12M %',
    'Intermediate Momentum': 'Int. Momentum %',
    'Return on invested capital %, Trailing 12 months': 'ROIC %'
}

def prepare_df(source_df, cols=None):
    if cols is None:
        cols = output_cols
    use_cols = [c for c in cols if c in source_df.columns]
    out = source_df[use_cols].copy()
    if 'Market capitalization' in out.columns:
        out['Market capitalization'] = (out['Market capitalization'] / 1e6).round(0)
    return out.rename(columns=header_map)

sheets = {}

# Full Ranking (ALL stocks including financials)
sheets['VC2 Full Ranking'] = prepare_df(df.sort_values('VC2'))

# TV50 sheets (filtered: no financials, positive P/E)
for name, countries in FILTERS.items():
    if countries:
        subset = cheapest_filtered[cheapest_filtered[COUNTRY_COL].isin(countries)].copy()
    else:
        subset = cheapest_filtered.copy()
    subset = subset.dropna(subset=[mom6_col]).sort_values(mom6_col, ascending=False)
    tv = subset.head(TV_SIZE)
    sheets[f'TV50 {name}'] = prepare_df(tv)
    print(f"  TV50 {name}: {len(subset)} no decil 1 → Top50={len(tv)}")

# MOM50 sheets (ALL stocks, intermediate momentum)
if has_momentum:
    for name, countries in FILTERS.items():
        if countries:
            subset = df[df[COUNTRY_COL].isin(countries)].copy()
        else:
            subset = df.copy()
        subset = subset.dropna(subset=['Intermediate Momentum'])
        subset = subset.sort_values('Intermediate Momentum', ascending=False)
        mom = subset.head(MOM_SIZE)
        sheets[f'MOM50 {name}'] = prepare_df(mom, cols=mom_output_cols)
        print(f"  MOM50 {name}: {len(subset)} com dados → Top50={len(mom)}")

# ============================================================================
# 6. WRITE EXCEL
# ============================================================================
print("A gerar Excel...")
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    for sheet_name, sheet_df in sheets.items():
        sheet_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

# ============================================================================
# 7. FORMAT EXCEL
# ============================================================================
print("A formatar...")
wb = load_workbook(out_path)

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill_vc2 = PatternFill('solid', fgColor='1F4E79')
header_fill_tv = PatternFill('solid', fgColor='2E75B6')
header_fill_mom = PatternFill('solid', fgColor='7B2D8B')
data_font = Font(name='Arial', size=10)
thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))

filter_colors = {
    'Global': PatternFill('solid', fgColor='F2F2F2'),
    'USA': PatternFill('solid', fgColor='DCE6F1'),
    'Europe': PatternFill('solid', fgColor='E2EFDA'),
    'Asia': PatternFill('solid', fgColor='FDE9D9'),
    'Emerging': PatternFill('solid', fgColor='F2DCDB'),
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    is_tv = sheet_name.startswith('TV')
    is_mom = sheet_name.startswith('MOM')
    fill = header_fill_mom if is_mom else (header_fill_tv if is_tv else header_fill_vc2)

    row_fill = None
    if is_tv or is_mom:
        for fname, fcolor in filter_colors.items():
            if fname in sheet_name:
                row_fill = fcolor
                break

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        hdr = ws.cell(row=1, column=col_idx).value or ""
        max_len = len(str(hdr))
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            if hdr in ('Ticker', 'Company', 'Country', 'Sector'):
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')
            if hdr == 'Mkt Cap ($M)':
                cell.number_format = '#,##0'
            elif hdr in ('P/E', 'P/S', 'P/B', 'P/CF', 'EV/EBITDA', 'VC2 Score'):
                cell.number_format = '#,##0.00'
            elif hdr == 'VC2 Decile':
                cell.number_format = '0'
            elif 'Yield' in hdr or 'Momentum' in hdr or 'ROIC' in hdr or 'Int.' in hdr:
                cell.number_format = '#,##0.00'
            val_len = len(str(cell.value or ""))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

wb.save(out_path)

# ============================================================================
# 8. GENERATE wizard-data.json (for website)
# ============================================================================
print("A gerar wizard-data.json...")

try:
    dt = datetime.strptime(file_date, "%Y-%m-%d")
    display_date = dt.strftime("%d/%m/%Y")
except:
    display_date = file_date

def safe_round(val, decimals=2):
    if pd.isna(val):
        return None
    try:
        return round(float(val), decimals)
    except:
        return None

def mktcap_millions(val):
    """Convert market cap to millions, rounded to integer."""
    if pd.isna(val):
        return None
    try:
        return int(round(float(val) / 1e6))
    except:
        return None

wizard_data = {
    "date": display_date,
    "value_cols": ["ticker", "name", "mktcap", "pe", "ps", "pb", "pcf", "eveb", "shy", "vc2"],
    "trend_cols": ["ticker", "name", "mktcap", "mom6m", "vc2"],
    "momentum_cols": ["ticker", "name", "mktcap", "mom12m", "mom6m", "intmom"],
    "value": {},
    "trend": {},
    "momentum": {}
}

# Value: cheapest by VC2 (filtered: no financials, positive P/E)
for name, countries in FILTERS.items():
    if countries:
        region_df = df_filtered[df_filtered[COUNTRY_COL].isin(countries)].copy()
    else:
        region_df = df_filtered.copy()
    top = region_df.nsmallest(JSON_TOP_N, 'VC2')
    entries = []
    for _, row in top.iterrows():
        entries.append([
            str(row['Symbol']),
            str(row.get('Description', '')),
            mktcap_millions(row.get('Market capitalization')),
            safe_round(row.get('Price to earnings ratio')),
            safe_round(row.get('Price to sales ratio')),
            safe_round(row.get('Price to book ratio')),
            safe_round(row.get('Price to cash flow ratio')),
            safe_round(row.get('Enterprise value to EBITDA ratio, Trailing 12 months')),
            safe_round(row.get('Shareholder Yield'), 1),
            safe_round(row.get('VC2'))
        ])
    wizard_data["value"][name] = entries

# Trend: VC2 Decile 1 (filtered), sorted by 6M momentum
for name, countries in FILTERS.items():
    if countries:
        region_cheap = cheapest_filtered[cheapest_filtered[COUNTRY_COL].isin(countries)].copy()
    else:
        region_cheap = cheapest_filtered.copy()
    region_cheap = region_cheap.dropna(subset=[mom6_col])
    top = region_cheap.nlargest(JSON_TOP_N, mom6_col)
    entries = []
    for _, row in top.iterrows():
        entries.append([
            str(row['Symbol']),
            str(row.get('Description', '')),
            mktcap_millions(row.get('Market capitalization')),
            safe_round(row.get(mom6_col), 1),
            safe_round(row.get('VC2'))
        ])
    wizard_data["trend"][name] = entries

# Momentum: ALL stocks (including financials), sorted by intermediate momentum
if has_momentum:
    for name, countries in FILTERS.items():
        if countries:
            region_df = df[df[COUNTRY_COL].isin(countries)].copy()
        else:
            region_df = df.copy()
        region_df = region_df.dropna(subset=['Intermediate Momentum'])
        top = region_df.nlargest(MOM_JSON_TOP_N, 'Intermediate Momentum')
        entries = []
        for _, row in top.iterrows():
            entries.append([
                str(row['Symbol']),
                str(row.get('Description', '')),
                mktcap_millions(row.get('Market capitalization')),
                safe_round(row.get(mom12_col), 1),
                safe_round(row.get(mom6_col), 1),
                safe_round(row.get('Intermediate Momentum'), 1)
            ])
        wizard_data["momentum"][name] = entries

json_path = os.path.join(SCRIPT_DIR, "wizard-data.json")
with open(json_path, 'w') as f:
    json.dump(wizard_data, f, separators=(',', ':'))

json_size = os.path.getsize(json_path)
print(f"  wizard-data.json: {json_size:,} bytes")

# ============================================================================
# 9. DONE
# ============================================================================
print(f"\n{'='*60}")
print(f"  CONCLUÍDO!")
print(f"{'='*60}")
print(f"  Ficheiro: {os.path.basename(out_path)}")
print(f"  JSON:     wizard-data.json ({json_size:,} bytes)")
print(f"  Sheets: {', '.join(sheets.keys())}")
print(f"{'='*60}")
